"""ExcelRepository — v1 persistence.

Four sheets, exactly as specified, plus columns that preserve the rule pack's
own vocabulary. The extra columns are why the Postgres migration is a genuine
drop-in: the schema is already normalised on assessment_id and the four sheets
map one-to-one onto four tables.

Writes are atomic: the workbook is built in full, written to a temporary file
and moved into place, so a crash never leaves a half-written report.
"""
from __future__ import annotations

import logging
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.assessment import Assessment
from app.models.performance import PerformanceMeasurement, PerformanceStatistics
from app.models.results import SecurityResult

log = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="1F3B4D")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)

#: Native result -> fill colour. Purely presentational.
_RESULT_FILL = {
    "PASS": PatternFill("solid", fgColor="D9EDE2"),
    "FAIL": PatternFill("solid", fgColor="F5DEDD"),
    "WARN": PatternFill("solid", fgColor="F7EFDA"),
    "INFORMATIONAL": PatternFill("solid", fgColor="DEEAF2"),
    "NOT_TESTABLE": PatternFill("solid", fgColor="EFEFEF"),
    "N/A": PatternFill("solid", fgColor="EFEFEF"),
}

SUMMARY_COLUMNS = [
    "assessment_id", "target_url", "timestamp", "status", "total_rules",
    "passed", "failed", "not_applicable", "unknown",
    "native_pass", "native_fail", "native_warn", "native_informational",
    "native_not_testable", "coverage_pct", "pack_version", "browser_version",
    "llm_model", "duration_seconds", "blocked_reason",
]
SECURITY_COLUMNS = [
    "assessment_id", "rule_id", "rule_name", "category", "result",
    "confidence", "evidence", "source_of_evidence", "timestamp",
    "native_result", "severity", "automation_tier", "observed_value",
    "unknown_reason", "evaluated_by", "source_file", "source_line",
]
PERFORMANCE_COLUMNS = [
    "assessment_id", "network_profile", "iteration", "dns_time", "tcp_time",
    "tls_time", "ttfb", "dom_content_loaded", "page_load_time",
    "transferred_bytes", "request_count", "failed_requests",
    "lcp", "cls", "inp", "redirect_count", "succeeded", "error",
]
STATISTICS_COLUMNS = [
    "assessment_id", "network_profile", "metric", "mean", "median", "min",
    "max", "stddev", "p95", "n", "success_rate", "failure_rate",
]


class ExcelRepository:
    """Buffers rows, then writes one workbook on commit()."""

    def __init__(self, path: str | Path):
        self.path = Path(path)
        self._assessment: Assessment | None = None
        self._security: list[SecurityResult] = []
        self._performance: list[PerformanceMeasurement] = []
        self._stats: list[PerformanceStatistics] = []
        self.summary_text = ""

    async def save_assessment(self, assessment: Assessment) -> None:
        self._assessment = assessment

    async def save_security_results(self, results: list[SecurityResult]) -> None:
        self._security.extend(results)

    async def save_performance_results(
            self, measurements: list[PerformanceMeasurement]) -> None:
        self._performance.extend(measurements)

    async def save_statistics(self, stats: list[PerformanceStatistics]) -> None:
        self._stats.extend(stats)

    # -- writing -----------------------------------------------------------
    def _sheet(self, wb: Workbook, title: str, columns: list[str], first: bool):
        ws = wb.active if first else wb.create_sheet()
        ws.title = title
        ws.append(columns)
        for i, _ in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=i)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        return ws

    @staticmethod
    def _autosize(ws, columns: list[str], cap: int = 60) -> None:
        for i, name in enumerate(columns, start=1):
            width = len(name) + 2
            for row in ws.iter_rows(min_row=2, min_col=i, max_col=i):
                v = row[0].value
                if v is not None:
                    width = max(width, min(len(str(v)) + 2, cap))
            ws.column_dimensions[get_column_letter(i)].width = min(width, cap)

    async def commit(self) -> str:
        """Write the workbook atomically. Returns the final path."""
        wb = Workbook()

        # --- Sheet 1: Assessment Summary
        ws = self._sheet(wb, "Assessment Summary", SUMMARY_COLUMNS, first=True)
        if a := self._assessment:
            t = a.tally
            ws.append([
                a.assessment_id, a.target_url,
                a.timestamp.replace(tzinfo=None), a.status.value, t.total,
                t.yes, t.no, t.not_applicable, t.unknown,
                t.native_pass, t.native_fail, t.native_warn,
                t.native_informational, t.native_not_testable,
                a.coverage_pct, a.pack_version, a.browser_version,
                a.llm_model, round(a.duration_seconds, 1), a.blocked_reason,
            ])
        self._autosize(ws, SUMMARY_COLUMNS)

        # --- Sheet 2: Security Results
        ws = self._sheet(wb, "Security Results", SECURITY_COLUMNS, first=False)
        for r in sorted(self._security, key=lambda x: x.rule_id):
            ws.append([
                r.assessment_id, r.rule_id, r.rule_name, r.category,
                r.result.value, round(r.confidence, 2), r.evidence,
                r.source_of_evidence, r.timestamp.replace(tzinfo=None),
                r.native_result.value, r.severity.value,
                r.automation_tier.value, r.observed_value, r.unknown_reason,
                r.evaluated_by, r.source_file, r.source_line,
            ])
            fill = _RESULT_FILL.get(r.native_result.value)
            if fill:
                ws.cell(row=ws.max_row,
                        column=SECURITY_COLUMNS.index("native_result") + 1).fill = fill
        for row in ws.iter_rows(min_row=2,
                                min_col=SECURITY_COLUMNS.index("evidence") + 1,
                                max_col=SECURITY_COLUMNS.index("evidence") + 1):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")
        self._autosize(ws, SECURITY_COLUMNS)

        # --- Sheet 3: Performance Results
        ws = self._sheet(wb, "Performance Results", PERFORMANCE_COLUMNS, first=False)
        for m in self._performance:
            ws.append([
                m.assessment_id, m.network_profile, m.iteration,
                m.dns_time, m.tcp_time, m.tls_time, m.ttfb,
                m.dom_content_loaded, m.page_load_time, m.transferred_bytes,
                m.request_count, m.failed_requests, m.lcp, m.cls, m.inp,
                m.redirect_count, m.succeeded, m.error,
            ])
        self._autosize(ws, PERFORMANCE_COLUMNS)

        # --- Sheet 4: Performance Statistics
        ws = self._sheet(wb, "Performance Statistics", STATISTICS_COLUMNS,
                         first=False)
        for s in self._stats:
            ws.append([
                s.assessment_id, s.network_profile, s.metric, s.mean, s.median,
                s.min, s.max, s.stddev, s.p95, s.n, s.success_rate,
                s.failure_rate,
            ])
        self._autosize(ws, STATISTICS_COLUMNS)

        # --- optional narrative sheet
        if self.summary_text:
            ws = wb.create_sheet("Summary")
            ws["A1"] = "Executive Summary"
            ws["A1"].font = Font(bold=True, size=12)
            ws["A2"] = self.summary_text
            ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions["A"].width = 120
            ws.row_dimensions[2].height = 220

        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.path.with_suffix(".tmp.xlsx")
        wb.save(tmp)
        os.replace(tmp, self.path)      # atomic on POSIX
        log.info("workbook written: %s", self.path)
        return str(self.path)

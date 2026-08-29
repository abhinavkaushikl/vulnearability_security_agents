"""Excel persistence: four sheets, round-trippable, atomic."""
from __future__ import annotations

from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from app.models.assessment import Assessment, AssessmentStatus
from app.models.performance import PerformanceMeasurement, PerformanceStatistics
from app.models.results import (ContractResult, NativeResult, ResultTally,
                                SecurityResult)
from app.models.rules import Automation, Severity
from app.repositories.base import AssessmentRepository
from app.repositories.excel import (PERFORMANCE_COLUMNS, SECURITY_COLUMNS,
                                    STATISTICS_COLUMNS, SUMMARY_COLUMNS,
                                    ExcelRepository)

REQUIRED_SUMMARY = ["assessment_id", "target_url", "timestamp", "status",
                    "total_rules", "passed", "failed", "not_applicable",
                    "unknown"]
REQUIRED_SECURITY = ["assessment_id", "rule_id", "rule_name", "category",
                     "result", "confidence", "evidence", "source_of_evidence",
                     "timestamp"]
REQUIRED_PERF = ["assessment_id", "network_profile", "iteration", "dns_time",
                 "tcp_time", "tls_time", "ttfb", "dom_content_loaded",
                 "page_load_time", "transferred_bytes", "request_count",
                 "failed_requests"]
REQUIRED_STATS = ["assessment_id", "network_profile", "metric", "mean",
                  "median", "min", "max", "stddev", "p95"]


def sample_result(rule_id="WEB-01", native=NativeResult.FAIL):
    return SecurityResult(
        assessment_id="a1", rule_id=rule_id, rule_name="Set CSP",
        category="WEB", result=ContractResult.NO, native_result=native,
        evidence="No Content-Security-Policy response header was observed.",
        observed_value="content-security-policy", source_of_evidence="HDR@/",
        confidence=0.95, severity=Severity.HIGH,
        automation_tier=Automation.HYBRID, source_file="Rules/03_x.md",
        source_line=19)


@pytest.fixture
def written(tmp_path):
    async def _write():
        repo = ExcelRepository(tmp_path / "out" / "assessment_results.xlsx")
        repo.summary_text = "An executive summary."
        await repo.save_assessment(Assessment(
            assessment_id="a1", target_url="https://x.test",
            timestamp=datetime.now(timezone.utc),
            status=AssessmentStatus.COMPLETED,
            tally=ResultTally(total=144, yes=8, no=3, not_applicable=1,
                              unknown=132, native_pass=8, native_fail=3,
                              native_not_testable=130),
            coverage_pct=7.6, pack_version="2026-08-28"))
        await repo.save_security_results([sample_result(),
                                          sample_result("WEB-02",
                                                        NativeResult.PASS)])
        await repo.save_performance_results([PerformanceMeasurement(
            assessment_id="a1", network_profile="3g", iteration=1,
            succeeded=True, ttfb=310.5, page_load_time=1820.0,
            request_count=12, failed_requests=0)])
        await repo.save_statistics([PerformanceStatistics(
            assessment_id="a1", network_profile="3g", metric="ttfb", n=3,
            mean=310.5, median=308.0, min=290.0, max=340.0, stddev=25.1,
            p95=338.0, success_rate=1.0)])
        return await repo.commit()

    import asyncio
    path = asyncio.run(_write())
    return load_workbook(path), path


def test_all_four_sheets_exist_in_order(written):
    wb, _ = written
    assert wb.sheetnames[:4] == ["Assessment Summary", "Security Results",
                                 "Performance Results", "Performance Statistics"]


@pytest.mark.parametrize("sheet,required,actual", [
    ("Assessment Summary", REQUIRED_SUMMARY, SUMMARY_COLUMNS),
    ("Security Results", REQUIRED_SECURITY, SECURITY_COLUMNS),
    ("Performance Results", REQUIRED_PERF, PERFORMANCE_COLUMNS),
    ("Performance Statistics", REQUIRED_STATS, STATISTICS_COLUMNS),
])
def test_every_specified_column_is_present(written, sheet, required, actual):
    wb, _ = written
    header = [c.value for c in wb[sheet][1]]
    assert header == actual
    for col in required:
        assert col in header, f"{sheet} is missing required column {col}"


def test_security_rows_round_trip(written):
    wb, _ = written
    ws = wb["Security Results"]
    header = [c.value for c in ws[1]]
    rows = {r[header.index("rule_id")]: r
            for r in ws.iter_rows(min_row=2, values_only=True)}
    assert set(rows) == {"WEB-01", "WEB-02"}
    web01 = rows["WEB-01"]
    assert web01[header.index("result")] == "NO"
    assert web01[header.index("native_result")] == "FAIL"
    assert web01[header.index("severity")] == "High"
    assert web01[header.index("source_line")] == 19


def test_both_vocabularies_are_persisted(written):
    """Sheet 2 must let a reader tell WARN from NOT_TESTABLE."""
    wb, _ = written
    header = [c.value for c in wb["Security Results"][1]]
    assert "result" in header and "native_result" in header
    assert "unknown_reason" in header


def test_summary_row_carries_coverage(written):
    wb, _ = written
    ws = wb["Assessment Summary"]
    header = [c.value for c in ws[1]]
    row = next(ws.iter_rows(min_row=2, values_only=True))
    assert row[header.index("total_rules")] == 144
    assert row[header.index("coverage_pct")] == 7.6
    assert row[header.index("status")] == "COMPLETED"


def test_commit_leaves_no_temporary_file(written):
    _, path = written
    from pathlib import Path
    assert Path(path).exists()
    assert not list(Path(path).parent.glob("*.tmp.xlsx"))


def test_repository_satisfies_the_protocol(tmp_path):
    assert isinstance(ExcelRepository(tmp_path / "x.xlsx"), AssessmentRepository)
